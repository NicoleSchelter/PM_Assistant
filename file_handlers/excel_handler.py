"""
Excel file handler for PM Analysis Tool.

This module provides the ExcelHandler class for processing Excel files (.xlsx, .xls)
including stakeholder registers, risk registers, and other structured project data.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from core.models import FileFormat, ValidationResult
from file_handlers.base_handler import BaseFileHandler
from utils.exceptions import DataExtractionError, FileProcessingError, ValidationError
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelHandler(BaseFileHandler):
    """
    Handler for Excel files (.xlsx, .xls) containing project management data.

    This handler can process various types of Excel files including:
    - Stakeholder registers
    - Risk registers
    - Work breakdown structures
    - Project schedules
    - Status reports

    Features:
    - Multi-sheet support
    - Structured table extraction
    - Error handling for corrupted files
    - Password-protected file detection
    - Data type inference and validation
    """

    def __init__(self):
        """Initialize the Excel handler."""
        super().__init__()
        self.supported_extensions = ["xlsx", "xls"]
        self.handler_name = "Excel Handler"

        # Common column patterns for different document types
        self.stakeholder_patterns = [
            "name",
            "stakeholder",
            "role",
            "contact",
            "email",
            "phone",
            "influence",
            "interest",
            "power",
            "attitude",
            "communication",
        ]

        self.risk_patterns = [
            "risk",
            "id",
            "description",
            "probability",
            "impact",
            "score",
            "status",
            "owner",
            "mitigation",
            "response",
            "category",
        ]

        self.deliverable_patterns = [
            "deliverable",
            "task",
            "wbs",
            "work",
            "package",
            "activity",
            "status",
            "progress",
            "start",
            "finish",
            "duration",
            "assigned",
        ]

    def can_handle(self, file_path: str) -> bool:
        """
        Check if this handler can process the given Excel file.

        Args:
            file_path (str): Path to the file to check

        Returns:
            bool: True if file is a supported Excel format, False otherwise
        """
        try:
            path = Path(file_path)
            extension = path.suffix.lower().lstrip(".")
            return extension in self.supported_extensions
        except Exception as e:
            logger.warning(f"Error checking file {file_path}: {e}")
            return False

    def extract_data(self, file_path: str) -> Dict[str, Any]:
        """
        Extract structured data from Excel file.

        Args:
            file_path (str): Path to the Excel file to process

        Returns:
            Dict[str, Any]: Extracted data including sheets, tables, and metadata

        Raises:
            FileProcessingError: If the file cannot be processed
            ValidationError: If the file format is invalid
        """
        try:
            logger.info(f"Extracting data from Excel file: {file_path}")

            # Validate file exists and is readable
            path = Path(file_path)
            if not path.exists():
                raise FileProcessingError(f"File not found: {file_path}")

            # Try to read with pandas first (handles both .xlsx and .xls)
            try:
                # Read all sheets
                excel_data = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
            except Exception as pandas_error:
                # Fallback to xlrd for older .xls files
                try:
                    excel_data = pd.read_excel(file_path, sheet_name=None, engine="xlrd")
                except Exception as xlrd_error:
                    raise FileProcessingError(
                        f"Failed to read Excel file with both engines. "
                        f"Pandas error: {pandas_error}. XLRd error: {xlrd_error}"
                    )

            # Extract metadata using openpyxl for .xlsx files
            metadata = self._extract_metadata(file_path)

            # Process each sheet
            processed_sheets = {}
            for sheet_name, df in excel_data.items():
                processed_sheets[sheet_name] = self._process_sheet(df, sheet_name)

            # Detect document type based on content
            document_type = self._detect_document_type(processed_sheets)

            # Extract structured tables
            structured_data = self._extract_structured_data(processed_sheets, document_type)

            result = {
                "file_path": str(path.absolute()),
                "file_name": path.name,
                "document_type": document_type,
                "metadata": metadata,
                "sheets": processed_sheets,
                "structured_data": structured_data,
                "summary": {
                    "total_sheets": len(processed_sheets),
                    "total_rows": sum(sheet["row_count"] for sheet in processed_sheets.values()),
                    "total_columns": sum(
                        sheet["column_count"] for sheet in processed_sheets.values()
                    ),
                    "has_data": any(sheet["has_data"] for sheet in processed_sheets.values()),
                },
            }

            logger.info(f"Successfully extracted data from {len(processed_sheets)} sheets")
            return result

        except Exception as e:
            error_msg = f"Failed to extract data from Excel file {file_path}: {str(e)}"
            logger.error(error_msg)
            raise FileProcessingError(error_msg) from e

    def validate_structure(self, file_path: str) -> ValidationResult:
        """
        Validate Excel file structure and content.

        Args:
            file_path (str): Path to the Excel file to validate

        Returns:
            ValidationResult: Validation result with success status and messages
        """
        result = ValidationResult(is_valid=True)

        try:
            path = Path(file_path)

            # Check file existence
            if not path.exists():
                result.add_error(f"File does not exist: {file_path}")
                return result

            # Check file extension
            if not self.can_handle(file_path):
                result.add_error(f"Unsupported file format: {path.suffix}")
                return result

            # Check file size (warn if too large)
            file_size_mb = path.stat().st_size / (1024 * 1024)
            if file_size_mb > 50:
                result.add_warning(f"Large file size: {file_size_mb:.1f}MB may slow processing")

            # Try to open and read the file
            try:
                # Test with openpyxl first
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()

                if not sheet_names:
                    result.add_error("Excel file contains no sheets")
                    return result

                # Test reading data with pandas
                excel_data = pd.read_excel(file_path, sheet_name=None, nrows=5)

                # Validate each sheet has some content
                empty_sheets = []
                for sheet_name, df in excel_data.items():
                    if df.empty:
                        empty_sheets.append(sheet_name)

                if empty_sheets:
                    if len(empty_sheets) == len(excel_data):
                        result.add_warning("All sheets in Excel file are empty")
                    else:
                        result.add_warning(f"Empty sheets found: {', '.join(empty_sheets)}")

                # Check for password protection
                try:
                    # Try to access workbook properties
                    workbook = load_workbook(file_path, read_only=True)
                    workbook.close()
                except InvalidFileException as e:
                    if "password" in str(e).lower():
                        result.add_error("File appears to be password protected")
                    else:
                        result.add_error(f"Invalid Excel file format: {e}")

            except PermissionError:
                result.add_error("Permission denied - file may be open in another application")
            except Exception as e:
                if "password" in str(e).lower():
                    result.add_error("File appears to be password protected")
                elif "corrupt" in str(e).lower():
                    result.add_error("File appears to be corrupted")
                else:
                    result.add_error(f"Failed to read Excel file: {str(e)}")

        except Exception as e:
            result.add_error(f"Validation error: {str(e)}")

        return result

    def _extract_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract metadata from Excel file using openpyxl."""
        metadata = {}

        try:
            # Only try openpyxl for .xlsx files
            if file_path.lower().endswith(".xlsx"):
                workbook = load_workbook(file_path, read_only=True)

                # Extract workbook properties
                props = workbook.properties
                if props:
                    metadata.update(
                        {
                            "title": props.title or "",
                            "author": props.creator or "",
                            "subject": props.subject or "",
                            "description": props.description or "",
                            "created": props.created.isoformat() if props.created else None,
                            "modified": props.modified.isoformat() if props.modified else None,
                            "last_modified_by": props.lastModifiedBy or "",
                        }
                    )

                # Extract sheet information
                metadata["sheets"] = []
                for sheet in workbook.worksheets:
                    sheet_info = {
                        "name": sheet.title,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "has_data": sheet.max_row > 1 or sheet.max_column > 1,
                    }
                    metadata["sheets"].append(sheet_info)

                workbook.close()
            else:
                # For .xls files, we can't extract detailed metadata
                metadata["extraction_error"] = "Metadata extraction not supported for .xls files"

        except Exception as e:
            logger.warning(f"Could not extract metadata from {file_path}: {e}")
            metadata["extraction_error"] = str(e)

        return metadata

    def _process_sheet(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """Process a single sheet's data."""
        # Clean the dataframe
        df_clean = df.dropna(how="all").dropna(axis=1, how="all")

        # Extract basic information
        sheet_info = {
            "name": sheet_name,
            "row_count": len(df_clean),
            "column_count": len(df_clean.columns),
            "has_data": not df_clean.empty,
            "columns": df_clean.columns.tolist(),
            "data_types": df_clean.dtypes.to_dict(),
            "sample_data": df_clean.head(3).to_dict("records") if not df_clean.empty else [],
        }

        # Detect potential data patterns
        if not df_clean.empty:
            sheet_info["patterns"] = self._detect_patterns(df_clean)

        return sheet_info

    def _detect_patterns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Detect data patterns in the dataframe."""
        patterns = {
            "likely_stakeholder_data": False,
            "likely_risk_data": False,
            "likely_deliverable_data": False,
            "has_dates": False,
            "has_status_column": False,
            "has_id_column": False,
        }

        # Convert column names to lowercase for pattern matching
        columns_lower = [str(col).lower() for col in df.columns]

        # Check for stakeholder patterns
        stakeholder_matches = sum(
            1
            for pattern in self.stakeholder_patterns
            if any(pattern in col for col in columns_lower)
        )
        patterns["likely_stakeholder_data"] = stakeholder_matches >= 3

        # Check for risk patterns
        risk_matches = sum(
            1 for pattern in self.risk_patterns if any(pattern in col for col in columns_lower)
        )
        patterns["likely_risk_data"] = risk_matches >= 3

        # Check for deliverable patterns
        deliverable_matches = sum(
            1
            for pattern in self.deliverable_patterns
            if any(pattern in col for col in columns_lower)
        )
        patterns["likely_deliverable_data"] = deliverable_matches >= 3

        # Check for common column types
        patterns["has_dates"] = any("date" in col or "time" in col for col in columns_lower)
        patterns["has_status_column"] = any("status" in col for col in columns_lower)
        patterns["has_id_column"] = any("id" in col or col.endswith("id") for col in columns_lower)

        return patterns

    def _detect_document_type(self, sheets: Dict[str, Dict[str, Any]]) -> str:
        """Detect the type of document based on sheet content."""
        # Count pattern matches across all sheets
        stakeholder_score = 0
        risk_score = 0
        deliverable_score = 0

        for sheet_info in sheets.values():
            if sheet_info.get("has_data", False):
                patterns = sheet_info.get("patterns", {})
                if patterns.get("likely_stakeholder_data", False):
                    stakeholder_score += 1
                if patterns.get("likely_risk_data", False):
                    risk_score += 1
                if patterns.get("likely_deliverable_data", False):
                    deliverable_score += 1

        # Determine document type based on highest score
        if stakeholder_score > risk_score and stakeholder_score > deliverable_score:
            return "stakeholder_register"
        elif risk_score > deliverable_score:
            return "risk_register"
        elif deliverable_score > 0:
            return "work_breakdown_structure"
        else:
            return "unknown"

    def _extract_structured_data(
        self, sheets: Dict[str, Dict[str, Any]], document_type: str
    ) -> Dict[str, Any]:
        """Extract structured data based on detected document type."""
        structured_data = {
            "document_type": document_type,
            "extracted_records": [],
            "field_mappings": {},
            "data_quality": {},
        }

        # Find the sheet with the most relevant data
        best_sheet = None
        best_score = 0

        for sheet_name, sheet_info in sheets.items():
            if not sheet_info.get("has_data", False):
                continue

            patterns = sheet_info.get("patterns", {})
            score = 0

            if document_type == "stakeholder_register" and patterns.get("likely_stakeholder_data"):
                score = 3
            elif document_type == "risk_register" and patterns.get("likely_risk_data"):
                score = 3
            elif document_type == "work_breakdown_structure" and patterns.get(
                "likely_deliverable_data"
            ):
                score = 3
            else:
                # General scoring based on data presence
                score = sheet_info.get("row_count", 0) / 10

            if score > best_score:
                best_score = score
                best_sheet = sheet_name

        if best_sheet:
            structured_data["primary_sheet"] = best_sheet
            structured_data["field_mappings"] = self._create_field_mappings(
                sheets[best_sheet], document_type
            )

        return structured_data

    def _create_field_mappings(
        self, sheet_info: Dict[str, Any], document_type: str
    ) -> Dict[str, str]:
        """Create field mappings for structured data extraction."""
        mappings = {}
        columns = [str(col).lower() for col in sheet_info.get("columns", [])]

        if document_type == "stakeholder_register":
            # Map stakeholder fields
            for col in columns:
                if "name" in col and "stakeholder" not in mappings:
                    mappings["stakeholder_name"] = col
                elif "role" in col and "role" not in mappings:
                    mappings["role"] = col
                elif "contact" in col or "email" in col:
                    mappings["contact"] = col
                elif "influence" in col:
                    mappings["influence"] = col
                elif "interest" in col:
                    mappings["interest"] = col

        elif document_type == "risk_register":
            # Map risk fields
            for col in columns:
                if "risk" in col and "description" not in mappings:
                    mappings["risk_description"] = col
                elif "probability" in col:
                    mappings["probability"] = col
                elif "impact" in col:
                    mappings["impact"] = col
                elif "status" in col:
                    mappings["status"] = col
                elif "owner" in col:
                    mappings["owner"] = col
                elif "mitigation" in col:
                    mappings["mitigation"] = col

        return mappings
