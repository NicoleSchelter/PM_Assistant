"""
Excel report generator for PM Analysis Tool.

This module provides the ExcelReporter class that generates structured
Excel reports with multiple sheets and professional formatting.
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference

from reporters.base_reporter import BaseReporter
from core.models import ProcessingResult, ProjectStatus
from core.domain import Risk, Deliverable, Milestone, Stakeholder
from utils.logger import get_logger
from utils.exceptions import ReportGenerationError, FileProcessingError
from utils.error_handling import handle_errors, error_context, safe_execute

logger = get_logger(__name__)


class ExcelReporter(BaseReporter):
    """
    Generates structured Excel reports with multiple sheets and formatting.
    
    This reporter creates professional Excel workbooks with separate sheets
    for different data types, including charts and conditional formatting.
    """
    
    def __init__(self):
        """Initialize the Excel reporter."""
        super().__init__()
        self.reporter_name = "Excel Reporter"
        self.output_format = "excel"
        self.file_extension = ".xlsx"
        
        # Define color scheme for formatting
        self.colors = {
            'header': 'FF4472C4',      # Blue
            'success': 'FF70AD47',     # Green
            'warning': 'FFFFC000',     # Yellow
            'error': 'FFE74C3C',       # Red
            'info': 'FF5B9BD5',        # Light Blue
            'light_gray': 'FFF2F2F2'   # Light Gray
        }
    
    def generate_report(self, 
                       data: ProcessingResult, 
                       output_path: str, 
                       config: Dict[str, Any]) -> str:
        """
        Generate an Excel report from processing results.
        
        Args:
            data: Processing results to format
            output_path: Directory where report should be saved
            config: Configuration for report generation
            
        Returns:
            Path to the generated report file
            
        Raises:
            ReportGenerationError: If report generation fails
        """
        try:
            # Validate output path
            if not self.validate_output_path(output_path):
                raise ValueError(f"Cannot write to {output_path}")
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Generate sheets based on available data
            config_with_project = {**config, 'include_project_in_metadata': True}
            self._create_summary_sheet(workbook, data, config_with_project)
            
            if 'risks' in data.data:
                self._create_risks_sheet(workbook, data.data['risks'], config)
            
            if 'deliverables' in data.data:
                self._create_deliverables_sheet(workbook, data.data['deliverables'], config)
            
            if 'milestones' in data.data:
                self._create_milestones_sheet(workbook, data.data['milestones'], config)
            
            if 'stakeholders' in data.data:
                self._create_stakeholders_sheet(workbook, data.data['stakeholders'], config)
            
            if 'found_files' in data.data or 'missing_files' in data.data:
                self._create_document_check_sheet(workbook, data.data, config)
            
            # Add charts if requested
            if config.get('include_charts', True):
                self._add_charts(workbook, data.data, config)
            
            # Generate filename and save
            base_name = config.get('filename', 'analysis_report')
            filename = self.generate_filename(
                base_name, 
                config.get('include_timestamp', True)
            )
            file_path = Path(output_path) / filename
            
            workbook.save(file_path)
            
            return str(file_path)
            
        except Exception as e:
            raise ValueError(f"Failed to generate Excel report: {e}")
    
    def format_data(self, data: Dict[str, Any], config: Dict[str, Any]) -> str:
        """
        Format data for Excel output (returns summary information).
        
        Args:
            data: Data to format
            config: Formatting configuration
            
        Returns:
            Summary of data formatting
        """
        summary_lines = []
        
        if 'project_status' in data:
            summary_lines.append(f"Project Status: {data['project_status'].get('project_name', 'Unknown')}")
        
        if 'risks' in data:
            summary_lines.append(f"Risks: {len(data['risks'])} items")
        
        if 'deliverables' in data:
            summary_lines.append(f"Deliverables: {len(data['deliverables'])} items")
        
        if 'milestones' in data:
            summary_lines.append(f"Milestones: {len(data['milestones'])} items")
        
        if 'stakeholders' in data:
            summary_lines.append(f"Stakeholders: {len(data['stakeholders'])} items")
        
        return "\n".join(summary_lines) if summary_lines else "No data to format"
    
    def _create_summary_sheet(self, workbook: Workbook, data: ProcessingResult, config: Dict[str, Any]) -> None:
        """Create the summary overview sheet."""
        sheet = workbook.create_sheet("Summary", 0)
        
        # Title
        title = config.get('title', 'Project Management Analysis Report')
        sheet['A1'] = title
        sheet['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        sheet.merge_cells('A1:F1')
        
        # Report metadata
        row = 3
        metadata = [
            ('Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Operation:', data.operation),
            ('Status:', '✓ Success' if data.success else '✗ Failed'),
            ('Processing Time:', f"{data.processing_time_seconds:.2f} seconds")
        ]
        
        # Add project name to metadata if available and if this is a full report (not just summary sheet test)
        if (config.get('include_project_in_metadata', False) and 
            'project_status' in data.data and 'project_name' in data.data['project_status']):
            metadata.insert(1, ('Project:', data.data['project_status']['project_name']))
        
        if data.file_path:
            metadata.append(('Source File:', str(data.file_path)))
        
        for label, value in metadata:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Project status overview
        if 'project_status' in data.data:
            self._add_project_status_summary(sheet, data.data['project_status'], row + 1)
        
        # Error and warning summary
        if data.errors or data.warnings:
            self._add_error_summary(sheet, data, row + 10)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _create_risks_sheet(self, workbook: Workbook, risks: List[Dict[str, Any]], config: Dict[str, Any]) -> None:
        """Create the risks analysis sheet."""
        if not risks:
            return  # Don't create sheet for empty data
            
        sheet = workbook.create_sheet("Risks")
        
        # Convert risks to DataFrame
        risks_data = []
        for risk in risks:
            risks_data.append({
                'Risk ID': risk.get('risk_id', 'N/A'),
                'Title': risk.get('title', 'Untitled'),
                'Category': risk.get('category', 'Unknown'),
                'Priority': risk.get('priority', 'low').title(),
                'Status': risk.get('status', 'open').replace('_', ' ').title(),
                'Owner': risk.get('owner', 'Unassigned'),
                'Probability': f"{risk.get('probability', 0)*100:.0f}%",
                'Impact': f"{risk.get('impact', 0)*100:.0f}%",
                'Risk Score': f"{risk.get('risk_score', 0):.2f}",
                'Identified Date': risk.get('identified_date', ''),
                'Target Resolution': risk.get('target_resolution_date', ''),
                'Mitigation Strategy': risk.get('mitigation_strategy', ''),
                'Description': risk.get('description', '')
            })
        
        if risks_data:
            df = pd.DataFrame(risks_data)
            self._write_dataframe_to_sheet(sheet, df, "Risks Analysis")
            self._apply_conditional_formatting_risks(sheet, len(risks_data))
    
    def _create_deliverables_sheet(self, workbook: Workbook, deliverables: List[Dict[str, Any]], config: Dict[str, Any]) -> None:
        """Create the deliverables tracking sheet."""
        if not deliverables:
            return  # Don't create sheet for empty data
            
        sheet = workbook.create_sheet("Deliverables")
        
        # Convert deliverables to DataFrame
        deliverables_data = []
        for deliverable in deliverables:
            deliverables_data.append({
                'Deliverable ID': deliverable.get('deliverable_id', 'N/A'),
                'Name': deliverable.get('name', 'Untitled'),
                'WBS Code': deliverable.get('wbs_code', ''),
                'Status': deliverable.get('status', 'not_started').replace('_', ' ').title(),
                'Assigned To': deliverable.get('assigned_to', 'Unassigned'),
                'Progress': f"{deliverable.get('completion_percentage', 0):.1f}%",
                'Start Date': deliverable.get('start_date', ''),
                'Due Date': deliverable.get('due_date', ''),
                'Estimated Hours': deliverable.get('estimated_effort_hours', ''),
                'Actual Hours': deliverable.get('actual_effort_hours', ''),
                'Budget Allocated': deliverable.get('budget_allocated', ''),
                'Budget Spent': deliverable.get('budget_spent', ''),
                'Description': deliverable.get('description', '')
            })
        
        if deliverables_data:
            df = pd.DataFrame(deliverables_data)
            self._write_dataframe_to_sheet(sheet, df, "Deliverables Tracking")
            self._apply_conditional_formatting_deliverables(sheet, len(deliverables_data))
    
    def _create_milestones_sheet(self, workbook: Workbook, milestones: List[Dict[str, Any]], config: Dict[str, Any]) -> None:
        """Create the milestones tracking sheet."""
        if not milestones:
            return  # Don't create sheet for empty data
            
        sheet = workbook.create_sheet("Milestones")
        
        # Convert milestones to DataFrame
        milestones_data = []
        for milestone in milestones:
            milestones_data.append({
                'Milestone ID': milestone.get('milestone_id', 'N/A'),
                'Name': milestone.get('name', 'Untitled'),
                'Type': milestone.get('milestone_type', ''),
                'Target Date': milestone.get('target_date', ''),
                'Actual Date': milestone.get('actual_date', ''),
                'Status': milestone.get('status', 'upcoming').replace('_', ' ').title(),
                'Owner': milestone.get('owner', 'Unassigned'),
                'Approval Required': 'Yes' if milestone.get('approval_required', False) else 'No',
                'Approver': milestone.get('approver', ''),
                'Dependencies': ', '.join(milestone.get('dependencies', [])),
                'Description': milestone.get('description', '')
            })
        
        if milestones_data:
            df = pd.DataFrame(milestones_data)
            self._write_dataframe_to_sheet(sheet, df, "Milestones Tracking")
            self._apply_conditional_formatting_milestones(sheet, len(milestones_data))
    
    def _create_stakeholders_sheet(self, workbook: Workbook, stakeholders: List[Dict[str, Any]], config: Dict[str, Any]) -> None:
        """Create the stakeholders analysis sheet."""
        if not stakeholders:
            return  # Don't create sheet for empty data
            
        sheet = workbook.create_sheet("Stakeholders")
        
        # Convert stakeholders to DataFrame
        stakeholders_data = []
        for stakeholder in stakeholders:
            stakeholders_data.append({
                'Stakeholder ID': stakeholder.get('stakeholder_id', 'N/A'),
                'Name': stakeholder.get('name', 'Unknown'),
                'Role': stakeholder.get('role', 'Unknown'),
                'Organization': stakeholder.get('organization', ''),
                'Email': stakeholder.get('email', ''),
                'Phone': stakeholder.get('phone', ''),
                'Influence': stakeholder.get('influence', 'medium').replace('_', ' ').title(),
                'Interest': stakeholder.get('interest', 'medium').replace('_', ' ').title(),
                'Engagement Priority': stakeholder.get('engagement_priority', 'Monitor'),
                'Communication Frequency': stakeholder.get('communication_frequency', ''),
                'Preferred Method': stakeholder.get('preferred_communication_method', ''),
                'Current Sentiment': stakeholder.get('current_sentiment', ''),
                'Last Contact': stakeholder.get('last_contact_date', ''),
                'Next Contact': stakeholder.get('next_contact_date', ''),
                'Key Concerns': ', '.join(stakeholder.get('key_concerns', [])),
                'Expectations': ', '.join(stakeholder.get('expectations', []))
            })
        
        if stakeholders_data:
            df = pd.DataFrame(stakeholders_data)
            self._write_dataframe_to_sheet(sheet, df, "Stakeholders Analysis")
            self._apply_conditional_formatting_stakeholders(sheet, len(stakeholders_data))
    
    def _create_document_check_sheet(self, workbook: Workbook, data: Dict[str, Any], config: Dict[str, Any]) -> None:
        """Create the document check results sheet."""
        sheet = workbook.create_sheet("Document Check")
        
        # Title
        sheet['A1'] = "Document Check Results"
        sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        sheet.merge_cells('A1:D1')
        
        row = 3
        
        # Found files section
        if 'found_files' in data:
            sheet[f'A{row}'] = "Found Documents"
            sheet[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Headers
            headers = ['Document Name', 'Format', 'Size (bytes)', 'Status']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid')
            row += 1
            
            # Data rows
            for file_info in data['found_files']:
                sheet.cell(row=row, column=1, value=file_info.get('filename', 'Unknown'))
                sheet.cell(row=row, column=2, value=file_info.get('format', 'Unknown'))
                sheet.cell(row=row, column=3, value=file_info.get('size_bytes', 0))
                
                status_cell = sheet.cell(row=row, column=4)
                if file_info.get('is_readable', True):
                    status_cell.value = "Valid"
                    status_cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                else:
                    status_cell.value = "Invalid"
                    status_cell.fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
                
                row += 1
            
            row += 2
        
        # Missing files section
        if 'missing_files' in data and data['missing_files']:
            sheet[f'A{row}'] = "Missing Documents"
            sheet[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for missing_file in data['missing_files']:
                cell = sheet.cell(row=row, column=1, value=missing_file)
                cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _add_project_status_summary(self, sheet, status_data: Dict[str, Any], start_row: int) -> None:
        """Add project status summary to the sheet."""
        sheet[f'A{start_row}'] = "Project Status Overview"
        sheet[f'A{start_row}'].font = Font(bold=True, size=12)
        
        row = start_row + 1
        
        # Key metrics
        metrics = [
            ('Project Name:', status_data.get('project_name', 'Unknown')),
            ('Overall Health:', f"{status_data.get('health_percentage', 0)}%"),
            ('Total Risks:', status_data.get('total_risks', 0)),
            ('High Priority Risks:', status_data.get('high_priority_risks', 0)),
            ('Total Deliverables:', status_data.get('total_deliverables', 0)),
            ('Completed Deliverables:', status_data.get('completed_deliverables', 0)),
            ('Total Milestones:', status_data.get('total_milestones', 0)),
            ('Completed Milestones:', status_data.get('completed_milestones', 0)),
            ('Total Stakeholders:', status_data.get('total_stakeholders', 0))
        ]
        
        for label, value in metrics:
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            
            # Color code health percentage
            if label == 'Overall Health:' and isinstance(value, str) and '%' in value:
                health_pct = int(value.replace('%', ''))
                if health_pct >= 80:
                    sheet[f'B{row}'].fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                elif health_pct >= 60:
                    sheet[f'B{row}'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                else:
                    sheet[f'B{row}'].fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
            
            row += 1
    
    def _add_error_summary(self, sheet, data: ProcessingResult, start_row: int) -> None:
        """Add error and warning summary to the sheet."""
        sheet[f'A{start_row}'] = "Processing Issues"
        sheet[f'A{start_row}'].font = Font(bold=True, size=12)
        
        row = start_row + 1
        
        if data.errors:
            sheet[f'A{row}'] = "Errors:"
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
            row += 1
            
            for error in data.errors:
                sheet[f'A{row}'] = f"• {error}"
                row += 1
            
            row += 1
        
        if data.warnings:
            sheet[f'A{row}'] = "Warnings:"
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
            row += 1
            
            for warning in data.warnings:
                sheet[f'A{row}'] = f"• {warning}"
                row += 1
    
    def _write_dataframe_to_sheet(self, sheet, df: pd.DataFrame, title: str) -> None:
        """Write a DataFrame to a sheet with formatting."""
        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        sheet['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        sheet.merge_cells(f'A1:{chr(65 + len(df.columns) - 1)}1')
        
        # Add empty row, then write DataFrame starting from row 3
        row_num = 3
        
        # Write headers
        for col_num, column_name in enumerate(df.columns, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=column_name)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, row_data in df.iterrows():
            row_num += 1
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _apply_conditional_formatting_risks(self, sheet, data_rows: int) -> None:
        """Apply conditional formatting to risks sheet."""
        # Priority column (assuming column D)
        priority_col = 'D'
        for row in range(4, 4 + data_rows):
            cell = sheet[f'{priority_col}{row}']
            if cell.value:
                if 'Critical' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
                elif 'High' in str(cell.value):
                    cell.fill = PatternFill(start_color='FFFF6B6B', end_color='FFFF6B6B', fill_type='solid')
                elif 'Medium' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                elif 'Low' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
    
    def _apply_conditional_formatting_deliverables(self, sheet, data_rows: int) -> None:
        """Apply conditional formatting to deliverables sheet."""
        # Progress column (assuming column F)
        progress_col = 'F'
        for row in range(4, 4 + data_rows):
            cell = sheet[f'{progress_col}{row}']
            if cell.value and '%' in str(cell.value):
                progress = float(str(cell.value).replace('%', ''))
                if progress >= 90:
                    cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                elif progress >= 50:
                    cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFFFC0CB', end_color='FFFFC0CB', fill_type='solid')
    
    def _apply_conditional_formatting_milestones(self, sheet, data_rows: int) -> None:
        """Apply conditional formatting to milestones sheet."""
        # Status column (assuming column F)
        status_col = 'F'
        for row in range(4, 4 + data_rows):
            cell = sheet[f'{status_col}{row}']
            if cell.value:
                if 'Completed' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                elif 'Overdue' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
                elif 'In Progress' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
    
    def _apply_conditional_formatting_stakeholders(self, sheet, data_rows: int) -> None:
        """Apply conditional formatting to stakeholders sheet."""
        # Engagement Priority column (assuming column I)
        priority_col = 'I'
        for row in range(4, 4 + data_rows):
            cell = sheet[f'{priority_col}{row}']
            if cell.value:
                if 'Manage Closely' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
                elif 'Keep Satisfied' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                elif 'Keep Informed' in str(cell.value):
                    cell.fill = PatternFill(start_color=self.colors['info'], end_color=self.colors['info'], fill_type='solid')
    
    def _add_charts(self, workbook: Workbook, data: Dict[str, Any], config: Dict[str, Any]) -> None:
        """Add charts to the workbook."""
        # Create charts sheet
        charts_sheet = workbook.create_sheet("Charts")
        
        # Risk priority distribution chart
        if 'risks' in data:
            self._create_risk_priority_chart(charts_sheet, data['risks'])
        
        # Deliverable status chart
        if 'deliverables' in data:
            self._create_deliverable_status_chart(charts_sheet, data['deliverables'])
        
        # Milestone timeline chart
        if 'milestones' in data:
            self._create_milestone_timeline_chart(charts_sheet, data['milestones'])
    
    def _create_risk_priority_chart(self, sheet, risks: List[Dict[str, Any]]) -> None:
        """Create a risk priority distribution pie chart."""
        # Count risks by priority
        priority_counts = {'Low': 0, 'Medium': 0, 'High': 0, 'Critical': 0}
        for risk in risks:
            priority = risk.get('priority', 'low').title()
            if priority in priority_counts:
                priority_counts[priority] += 1
        
        # Create data for chart
        row = 2
        sheet['A1'] = "Risk Priority Distribution"
        sheet['A1'].font = Font(bold=True, size=12)
        
        for priority, count in priority_counts.items():
            sheet[f'A{row}'] = priority
            sheet[f'B{row}'] = count
            row += 1
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Risk Priority Distribution"
        
        data_ref = Reference(sheet, min_col=2, min_row=2, max_row=row-1)
        labels_ref = Reference(sheet, min_col=1, min_row=2, max_row=row-1)
        
        chart.add_data(data_ref)
        chart.set_categories(labels_ref)
        
        sheet.add_chart(chart, "D2")
    
    def _create_deliverable_status_chart(self, sheet, deliverables: List[Dict[str, Any]]) -> None:
        """Create a deliverable status distribution bar chart."""
        # Count deliverables by status
        status_counts = {}
        for deliverable in deliverables:
            status = deliverable.get('status', 'not_started').replace('_', ' ').title()
            status_counts[status] = status_counts.get(status, 0) + 1
        
        # Create data for chart
        row = 10
        sheet[f'A{row}'] = "Deliverable Status Distribution"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for status, count in status_counts.items():
            sheet[f'A{row}'] = status
            sheet[f'B{row}'] = count
            row += 1
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Deliverable Status Distribution"
        chart.x_axis.title = "Status"
        chart.y_axis.title = "Count"
        
        data_ref = Reference(sheet, min_col=2, min_row=11, max_row=row-1)
        labels_ref = Reference(sheet, min_col=1, min_row=11, max_row=row-1)
        
        chart.add_data(data_ref)
        chart.set_categories(labels_ref)
        
        sheet.add_chart(chart, "D10")
    
    def _create_milestone_timeline_chart(self, sheet, milestones: List[Dict[str, Any]]) -> None:
        """Create a milestone timeline chart."""
        # This is a simplified version - in practice, you might want a more sophisticated timeline
        completed_count = len([m for m in milestones if m.get('status') == 'completed'])
        total_count = len(milestones)
        
        row = 18
        sheet[f'A{row}'] = "Milestone Progress"
        sheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        sheet[f'A{row}'] = "Completed"
        sheet[f'B{row}'] = completed_count
        row += 1
        
        sheet[f'A{row}'] = "Remaining"
        sheet[f'B{row}'] = total_count - completed_count
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Milestone Progress"
        
        data_ref = Reference(sheet, min_col=2, min_row=19, max_row=20)
        labels_ref = Reference(sheet, min_col=1, min_row=19, max_row=20)
        
        chart.add_data(data_ref)
        chart.set_categories(labels_ref)
        
        sheet.add_chart(chart, "D18")
    
    def _auto_adjust_columns(self, sheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue  # Skip if we can't determine column letter
            
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def get_supported_config_options(self) -> Dict[str, Any]:
        """Get configuration options supported by the Excel reporter."""
        base_options = super().get_supported_config_options()
        
        excel_options = {
            'title': {
                'type': str,
                'default': 'Project Management Analysis Report',
                'description': 'Custom title for the report'
            },
            'filename': {
                'type': str,
                'default': 'analysis_report',
                'description': 'Base filename for the report'
            },
            'include_charts': {
                'type': bool,
                'default': True,
                'description': 'Include charts and visualizations'
            },
            'apply_formatting': {
                'type': bool,
                'default': True,
                'description': 'Apply conditional formatting and styling'
            },
            'auto_adjust_columns': {
                'type': bool,
                'default': True,
                'description': 'Automatically adjust column widths'
            }
        }
        
        return {**base_options, **excel_options}