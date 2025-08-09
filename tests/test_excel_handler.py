"""
Unit tests for Excel file handler.

This module contains comprehensive tests for the ExcelHandler class,
including tests with sample Excel files for stakeholder and risk registers.
"""

import pytest
import pandas as pd
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import tempfile
import os
from openpyxl import Workbook

from file_handlers.excel_handler import ExcelHandler
from core.models import ValidationResult
from utils.exceptions import FileProcessingError, ValidationError


class TestExcelHandler:
    """Test cases for ExcelHandler class."""
    
    @pytest.fixture
    def handler(self):
        """Create an ExcelHandler instance for testing."""
        return ExcelHandler()
    
    @pytest.fixture
    def sample_stakeholder_data(self):
        """Sample stakeholder register data."""
        return pd.DataFrame({
            'Stakeholder Name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
            'Role': ['Project Manager', 'Business Analyst', 'Developer'],
            'Contact Email': ['john@company.com', 'jane@company.com', 'bob@company.com'],
            'Influence': ['High', 'Medium', 'Low'],
            'Interest': ['High', 'High', 'Medium'],
            'Communication Preference': ['Email', 'Phone', 'Email']
        })
    
    @pytest.fixture
    def sample_risk_data(self):
        """Sample risk register data."""
        return pd.DataFrame({
            'Risk ID': ['R001', 'R002', 'R003'],
            'Risk Description': [
                'Budget overrun due to scope creep',
                'Key team member unavailable',
                'Technology integration issues'
            ],
            'Probability': ['Medium', 'Low', 'High'],
            'Impact': ['High', 'Medium', 'Medium'],
            'Status': ['Open', 'Mitigated', 'Open'],
            'Owner': ['PM', 'HR', 'Tech Lead'],
            'Mitigation Strategy': [
                'Regular scope reviews',
                'Cross-training team members',
                'Proof of concept development'
            ]
        })
    
    @pytest.fixture
    def temp_excel_file(self, sample_stakeholder_data):
        """Create a temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            sample_stakeholder_data.to_excel(tmp.name, index=False, sheet_name='Stakeholders')
            yield tmp.name
        os.unlink(tmp.name)
    
    @pytest.fixture
    def temp_multi_sheet_excel(self, sample_stakeholder_data, sample_risk_data):
        """Create a temporary multi-sheet Excel file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                sample_stakeholder_data.to_excel(writer, sheet_name='Stakeholders', index=False)
                sample_risk_data.to_excel(writer, sheet_name='Risks', index=False)
            yield tmp.name
        os.unlink(tmp.name)
    
    @pytest.fixture
    def temp_empty_excel(self):
        """Create a temporary empty Excel file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            workbook = Workbook()
            workbook.save(tmp.name)
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_initialization(self, handler):
        """Test ExcelHandler initialization."""
        assert handler.handler_name == "Excel Handler"
        assert handler.supported_extensions == ['xlsx', 'xls']
        assert 'name' in handler.stakeholder_patterns
        assert 'risk' in handler.risk_patterns
        assert 'deliverable' in handler.deliverable_patterns
    
    def test_can_handle_xlsx_file(self, handler):
        """Test can_handle method with .xlsx file."""
        assert handler.can_handle('test.xlsx') is True
        assert handler.can_handle('test.XLSX') is True
        assert handler.can_handle('/path/to/file.xlsx') is True
    
    def test_can_handle_xls_file(self, handler):
        """Test can_handle method with .xls file."""
        assert handler.can_handle('test.xls') is True
        assert handler.can_handle('test.XLS') is True
        assert handler.can_handle('/path/to/file.xls') is True
    
    def test_can_handle_unsupported_file(self, handler):
        """Test can_handle method with unsupported file types."""
        assert handler.can_handle('test.txt') is False
        assert handler.can_handle('test.pdf') is False
        assert handler.can_handle('test.docx') is False
        assert handler.can_handle('test') is False
    
    def test_can_handle_invalid_path(self, handler):
        """Test can_handle method with invalid path."""
        assert handler.can_handle('') is False
        assert handler.can_handle(None) is False
    
    def test_extract_data_stakeholder_register(self, handler, temp_excel_file):
        """Test data extraction from stakeholder register."""
        result = handler.extract_data(temp_excel_file)
        
        assert result['file_name'].endswith('.xlsx')
        assert result['document_type'] == 'stakeholder_register'
        assert 'Stakeholders' in result['sheets']
        assert result['summary']['total_sheets'] == 1
        assert result['summary']['has_data'] is True
        
        # Check sheet data
        sheet_data = result['sheets']['Stakeholders']
        assert sheet_data['has_data'] is True
        assert sheet_data['row_count'] == 3
        assert 'Stakeholder Name' in sheet_data['columns']
        assert sheet_data['patterns']['likely_stakeholder_data'] is True
    
    def test_extract_data_multi_sheet(self, handler, temp_multi_sheet_excel):
        """Test data extraction from multi-sheet Excel file."""
        result = handler.extract_data(temp_multi_sheet_excel)
        
        assert result['summary']['total_sheets'] == 2
        assert 'Stakeholders' in result['sheets']
        assert 'Risks' in result['sheets']
        
        # Check stakeholder sheet
        stakeholder_sheet = result['sheets']['Stakeholders']
        assert stakeholder_sheet['patterns']['likely_stakeholder_data'] is True
        
        # Check risk sheet
        risk_sheet = result['sheets']['Risks']
        assert risk_sheet['patterns']['likely_risk_data'] is True
    
    def test_extract_data_nonexistent_file(self, handler):
        """Test data extraction with nonexistent file."""
        with pytest.raises(FileProcessingError, match="File not found"):
            handler.extract_data('nonexistent.xlsx')
    
    @patch('pandas.read_excel')
    def test_extract_data_corrupted_file(self, mock_read_excel, handler, temp_excel_file):
        """Test data extraction with corrupted file."""
        mock_read_excel.side_effect = Exception("File is corrupted")
        
        with pytest.raises(FileProcessingError, match="Failed to extract data"):
            handler.extract_data(temp_excel_file)
    
    def test_validate_structure_valid_file(self, handler, temp_excel_file):
        """Test structure validation with valid Excel file."""
        result = handler.validate_structure(temp_excel_file)
        
        assert result.is_valid is True
        assert len(result.errors) == 0
    
    def test_validate_structure_nonexistent_file(self, handler):
        """Test structure validation with nonexistent file."""
        result = handler.validate_structure('nonexistent.xlsx')
        
        assert result.is_valid is False
        assert any('does not exist' in error for error in result.errors)
    
    def test_validate_structure_unsupported_format(self, handler):
        """Test structure validation with unsupported file format."""
        # Create a temporary text file to test unsupported format
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as tmp:
            tmp.write(b'test content')
            tmp.flush()
            
            result = handler.validate_structure(tmp.name)
            
            assert result.is_valid is False
            assert any('Unsupported file format' in error for error in result.errors)
        
        os.unlink(tmp.name)
    
    def test_validate_structure_empty_file(self, handler, temp_empty_excel):
        """Test structure validation with empty Excel file."""
        result = handler.validate_structure(temp_empty_excel)
        
        # Should be valid but may have warnings about empty sheets
        assert result.is_valid is True
        # Should have warnings about empty sheets
        assert len(result.warnings) > 0
    
    @patch('openpyxl.load_workbook')
    def test_validate_structure_password_protected(self, mock_load_workbook, handler, temp_excel_file):
        """Test structure validation with password-protected file."""
        mock_load_workbook.side_effect = Exception("password protected")
        
        result = handler.validate_structure(temp_excel_file)
        
        assert result.is_valid is False
        assert any('password protected' in error.lower() for error in result.errors)
    
    @patch('openpyxl.load_workbook')
    def test_validate_structure_permission_denied(self, mock_load_workbook, handler, temp_excel_file):
        """Test structure validation with permission denied."""
        mock_load_workbook.side_effect = PermissionError("Permission denied")
        
        result = handler.validate_structure(temp_excel_file)
        
        assert result.is_valid is False
        assert any('Permission denied' in error for error in result.errors)
    
    def test_detect_patterns_stakeholder_data(self, handler, sample_stakeholder_data):
        """Test pattern detection for stakeholder data."""
        patterns = handler._detect_patterns(sample_stakeholder_data)
        
        assert patterns['likely_stakeholder_data'] is True
        assert patterns['likely_risk_data'] is False
        assert patterns['has_id_column'] is False
    
    def test_detect_patterns_risk_data(self, handler, sample_risk_data):
        """Test pattern detection for risk data."""
        patterns = handler._detect_patterns(sample_risk_data)
        
        assert patterns['likely_risk_data'] is True
        assert patterns['likely_stakeholder_data'] is False
        assert patterns['has_id_column'] is True
        assert patterns['has_status_column'] is True
    
    def test_detect_patterns_empty_dataframe(self, handler):
        """Test pattern detection with empty dataframe."""
        empty_df = pd.DataFrame()
        patterns = handler._detect_patterns(empty_df)
        
        assert patterns['likely_stakeholder_data'] is False
        assert patterns['likely_risk_data'] is False
        assert patterns['likely_deliverable_data'] is False
    
    def test_detect_document_type_stakeholder(self, handler):
        """Test document type detection for stakeholder register."""
        sheets = {
            'Sheet1': {
                'has_data': True,
                'patterns': {
                    'likely_stakeholder_data': True,
                    'likely_risk_data': False,
                    'likely_deliverable_data': False
                }
            }
        }
        
        doc_type = handler._detect_document_type(sheets)
        assert doc_type == 'stakeholder_register'
    
    def test_detect_document_type_risk(self, handler):
        """Test document type detection for risk register."""
        sheets = {
            'Sheet1': {
                'has_data': True,
                'patterns': {
                    'likely_stakeholder_data': False,
                    'likely_risk_data': True,
                    'likely_deliverable_data': False
                }
            }
        }
        
        doc_type = handler._detect_document_type(sheets)
        assert doc_type == 'risk_register'
    
    def test_detect_document_type_unknown(self, handler):
        """Test document type detection for unknown document."""
        sheets = {
            'Sheet1': {
                'has_data': True,
                'patterns': {
                    'likely_stakeholder_data': False,
                    'likely_risk_data': False,
                    'likely_deliverable_data': False
                }
            }
        }
        
        doc_type = handler._detect_document_type(sheets)
        assert doc_type == 'unknown'
    
    def test_process_sheet_with_data(self, handler, sample_stakeholder_data):
        """Test sheet processing with data."""
        sheet_info = handler._process_sheet(sample_stakeholder_data, 'Test Sheet')
        
        assert sheet_info['name'] == 'Test Sheet'
        assert sheet_info['has_data'] is True
        assert sheet_info['row_count'] == 3
        assert sheet_info['column_count'] == 6
        assert len(sheet_info['sample_data']) == 3
        assert 'patterns' in sheet_info
    
    def test_process_sheet_empty(self, handler):
        """Test sheet processing with empty dataframe."""
        empty_df = pd.DataFrame()
        sheet_info = handler._process_sheet(empty_df, 'Empty Sheet')
        
        assert sheet_info['name'] == 'Empty Sheet'
        assert sheet_info['has_data'] is False
        assert sheet_info['row_count'] == 0
        assert sheet_info['column_count'] == 0
        assert sheet_info['sample_data'] == []
    
    def test_create_field_mappings_stakeholder(self, handler):
        """Test field mapping creation for stakeholder register."""
        sheet_info = {
            'columns': ['Stakeholder Name', 'Role', 'Contact Email', 'Influence', 'Interest']
        }
        
        mappings = handler._create_field_mappings(sheet_info, 'stakeholder_register')
        
        assert 'stakeholder_name' in mappings
        assert 'role' in mappings
        assert 'contact' in mappings
        assert 'influence' in mappings
        assert 'interest' in mappings
    
    def test_create_field_mappings_risk(self, handler):
        """Test field mapping creation for risk register."""
        sheet_info = {
            'columns': ['Risk Description', 'Probability', 'Impact', 'Status', 'Owner']
        }
        
        mappings = handler._create_field_mappings(sheet_info, 'risk_register')
        
        assert 'risk_description' in mappings
        assert 'probability' in mappings
        assert 'impact' in mappings
        assert 'status' in mappings
        assert 'owner' in mappings
    
    def test_extract_metadata_xlsx(self, handler, temp_excel_file):
        """Test metadata extraction from .xlsx file."""
        metadata = handler._extract_metadata(temp_excel_file)
        
        assert 'sheets' in metadata
        assert len(metadata['sheets']) > 0
        assert metadata['sheets'][0]['name'] == 'Stakeholders'
    
    def test_extract_metadata_xls(self, handler):
        """Test metadata extraction from .xls file (should handle gracefully)."""
        # Create a mock .xls file path
        fake_xls_path = 'test.xls'
        metadata = handler._extract_metadata(fake_xls_path)
        
        # Should handle the error gracefully
        assert 'extraction_error' in metadata
    
    def test_get_supported_extensions(self, handler):
        """Test getting supported extensions."""
        extensions = handler.get_supported_extensions()
        
        assert 'xlsx' in extensions
        assert 'xls' in extensions
        assert len(extensions) == 2
    
    def test_string_representation(self, handler):
        """Test string representation of handler."""
        str_repr = str(handler)
        assert 'Excel Handler' in str_repr
        assert 'xlsx' in str_repr
        assert 'xls' in str_repr
    
    @patch('file_handlers.excel_handler.logger')
    def test_logging_on_success(self, mock_logger, handler, temp_excel_file):
        """Test that successful operations are logged."""
        handler.extract_data(temp_excel_file)
        
        # Check that info logs were called
        mock_logger.info.assert_called()
    
    @patch('file_handlers.excel_handler.logger')
    def test_logging_on_error(self, mock_logger, handler):
        """Test that errors are logged."""
        with pytest.raises(FileProcessingError):
            handler.extract_data('nonexistent.xlsx')
        
        # Check that error logs were called
        mock_logger.error.assert_called()
    
    def test_large_file_warning(self, handler):
        """Test warning for large files."""
        # Create a mock large file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            # Write some data to make it appear large
            tmp.write(b'0' * (60 * 1024 * 1024))  # 60MB
            tmp.flush()
            
            # Mock the file size check
            with patch('pathlib.Path.stat') as mock_stat:
                mock_stat.return_value.st_size = 60 * 1024 * 1024
                
                result = handler.validate_structure(tmp.name)
                
                # Should have a warning about large file size
                assert any('Large file size' in warning for warning in result.warnings)
        
        os.unlink(tmp.name)


class TestExcelHandlerIntegration:
    """Integration tests for ExcelHandler with real Excel files."""
    
    @pytest.fixture
    def complex_stakeholder_file(self):
        """Create a complex stakeholder register file."""
        data = pd.DataFrame({
            'ID': range(1, 11),
            'Stakeholder Name': [f'Person {i}' for i in range(1, 11)],
            'Organization': ['Company A', 'Company B'] * 5,
            'Role/Title': ['Manager', 'Analyst', 'Developer', 'Tester', 'Admin'] * 2,
            'Contact Email': [f'person{i}@company.com' for i in range(1, 11)],
            'Phone': [f'555-000{i:02d}' for i in range(1, 11)],
            'Influence Level': ['High', 'Medium', 'Low'] * 3 + ['High'],
            'Interest Level': ['High', 'Medium', 'Low', 'Medium', 'High'] * 2,
            'Attitude': ['Supporter', 'Neutral', 'Opponent'] * 3 + ['Supporter'],
            'Communication Preference': ['Email', 'Phone', 'Meeting'] * 3 + ['Email'],
            'Notes': [f'Notes for person {i}' for i in range(1, 11)]
        })
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            data.to_excel(tmp.name, index=False, sheet_name='Stakeholder Register')
            yield tmp.name
        os.unlink(tmp.name)
    
    @pytest.fixture
    def complex_risk_file(self):
        """Create a complex risk register file."""
        data = pd.DataFrame({
            'Risk ID': [f'R{i:03d}' for i in range(1, 16)],
            'Risk Category': ['Technical', 'Schedule', 'Budget', 'Resource', 'External'] * 3,
            'Risk Description': [f'Risk description {i}' for i in range(1, 16)],
            'Probability': ['High', 'Medium', 'Low'] * 5,
            'Impact': ['High', 'Medium', 'Low'] * 5,
            'Risk Score': [9, 6, 3, 6, 3, 3, 9, 6, 3, 6, 3, 3, 9, 6, 3],
            'Status': ['Open', 'Mitigated', 'Closed'] * 5,
            'Owner': [f'Owner {i}' for i in range(1, 16)],
            'Mitigation Strategy': [f'Mitigation strategy {i}' for i in range(1, 16)],
            'Target Date': pd.date_range('2024-01-01', periods=15, freq='W'),
            'Last Updated': pd.date_range('2023-12-01', periods=15, freq='D')
        })
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            data.to_excel(tmp.name, index=False, sheet_name='Risk Register')
            yield tmp.name
        os.unlink(tmp.name)
    
    def test_complex_stakeholder_extraction(self, complex_stakeholder_file):
        """Test extraction from complex stakeholder register."""
        handler = ExcelHandler()
        result = handler.extract_data(complex_stakeholder_file)
        
        assert result['document_type'] == 'stakeholder_register'
        assert result['summary']['total_rows'] == 10
        
        # Check that all expected columns are present
        sheet_data = result['sheets']['Stakeholder Register']
        expected_columns = ['ID', 'Stakeholder Name', 'Role/Title', 'Contact Email', 
                          'Influence Level', 'Interest Level']
        for col in expected_columns:
            assert col in sheet_data['columns']
        
        # Check field mappings
        structured_data = result['structured_data']
        assert 'field_mappings' in structured_data
        assert 'primary_sheet' in structured_data
    
    def test_complex_risk_extraction(self, complex_risk_file):
        """Test extraction from complex risk register."""
        handler = ExcelHandler()
        result = handler.extract_data(complex_risk_file)
        
        assert result['document_type'] == 'risk_register'
        assert result['summary']['total_rows'] == 15
        
        # Check that all expected columns are present
        sheet_data = result['sheets']['Risk Register']
        expected_columns = ['Risk ID', 'Risk Description', 'Probability', 'Impact', 
                          'Status', 'Owner']
        for col in expected_columns:
            assert col in sheet_data['columns']
        
        # Check patterns detection
        patterns = sheet_data['patterns']
        assert patterns['likely_risk_data'] is True
        assert patterns['has_dates'] is True
        assert patterns['has_status_column'] is True
        assert patterns['has_id_column'] is True