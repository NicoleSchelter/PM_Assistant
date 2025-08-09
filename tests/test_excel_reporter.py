"""
Unit tests for ExcelReporter class.
"""

import pytest
from pathlib import Path
from datetime import datetime, date
from unittest.mock import patch, MagicMock
import tempfile
import os
import pandas as pd
from openpyxl import load_workbook

from reporters.excel_reporter import ExcelReporter
from core.models import ProcessingResult, ProjectStatus
from core.domain import Risk, Deliverable, Milestone, Stakeholder, RiskPriority, RiskStatus


class TestExcelReporter:
    """Test cases for ExcelReporter class."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.reporter = ExcelReporter()
        self.temp_dir = tempfile.mkdtemp()
    
    def teardown_method(self):
        """Clean up test fixtures."""
        # Clean up temp directory
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
    
    def test_init(self):
        """Test reporter initialization."""
        assert self.reporter.reporter_name == "Excel Reporter"
        assert self.reporter.output_format == "excel"
        assert self.reporter.file_extension == ".xlsx"
        assert 'header' in self.reporter.colors
        assert 'success' in self.reporter.colors
        assert 'warning' in self.reporter.colors
        assert 'error' in self.reporter.colors
    
    def test_generate_report_success(self):
        """Test successful report generation."""
        # Create test data
        processing_result = ProcessingResult(
            success=True,
            operation="test_operation",
            data={
                'project_status': {
                    'project_name': 'Test Project',
                    'overall_health_score': 0.85,
                    'health_percentage': 85,
                    'total_risks': 5,
                    'high_priority_risks': 2,
                    'total_deliverables': 10,
                    'completed_deliverables': 7,
                    'total_milestones': 4,
                    'completed_milestones': 2,
                    'total_stakeholders': 8
                },
                'risks': [
                    {
                        'risk_id': 'RISK-001',
                        'title': 'Test Risk',
                        'category': 'Technical',
                        'priority': 'high',
                        'status': 'open',
                        'owner': 'John Doe',
                        'probability': 0.7,
                        'impact': 0.8,
                        'risk_score': 0.56,
                        'identified_date': '2024-01-15',
                        'mitigation_strategy': 'Test mitigation',
                        'description': 'Test risk description'
                    }
                ]
            },
            processing_time_seconds=2.5
        )
        
        config = {
            'title': 'Test Excel Report',
            'filename': 'test_excel_report',
            'include_timestamp': False,
            'include_charts': True
        }
        
        # Generate report
        report_path = self.reporter.generate_report(processing_result, self.temp_dir, config)
        
        # Verify file was created
        assert os.path.exists(report_path)
        assert report_path.endswith('.xlsx')
        
        # Verify workbook structure
        workbook = load_workbook(report_path)
        sheet_names = workbook.sheetnames
        
        assert 'Summary' in sheet_names
        assert 'Risks' in sheet_names
        assert 'Charts' in sheet_names
        
        # Verify summary sheet content
        summary_sheet = workbook['Summary']
        assert summary_sheet['A1'].value == 'Test Excel Report'
        assert 'Test Project' in str(summary_sheet['B4'].value)  # Project name should be in metadata
    
    def test_generate_report_with_all_data_types(self):
        """Test report generation with all data types."""
        processing_result = ProcessingResult(
            success=True,
            operation="comprehensive_analysis",
            data={
                'project_status': {
                    'project_name': 'Comprehensive Project',
                    'health_percentage': 75
                },
                'risks': [{'risk_id': 'R001', 'title': 'Risk 1', 'priority': 'medium'}],
                'deliverables': [{'deliverable_id': 'D001', 'name': 'Deliverable 1', 'status': 'in_progress'}],
                'milestones': [{'milestone_id': 'M001', 'name': 'Milestone 1', 'status': 'upcoming'}],
                'stakeholders': [{'stakeholder_id': 'S001', 'name': 'Stakeholder 1', 'role': 'Sponsor'}],
                'found_files': [{'filename': 'test.xlsx', 'format': 'excel', 'is_readable': True}],
                'missing_files': ['missing_doc.md']
            },
            processing_time_seconds=3.0
        )
        
        config = {'filename': 'comprehensive_report', 'include_timestamp': False}
        
        report_path = self.reporter.generate_report(processing_result, self.temp_dir, config)
        
        # Verify all sheets were created
        workbook = load_workbook(report_path)
        sheet_names = workbook.sheetnames
        
        expected_sheets = ['Summary', 'Risks', 'Deliverables', 'Milestones', 'Stakeholders', 'Document Check', 'Charts']
        for sheet in expected_sheets:
            assert sheet in sheet_names
    
    def test_generate_report_with_timestamp(self):
        """Test report generation with timestamp in filename."""
        processing_result = ProcessingResult(
            success=True,
            operation="test_operation",
            data={},
            processing_time_seconds=1.0
        )
        
        config = {
            'filename': 'timestamped_excel_report',
            'include_timestamp': True
        }
        
        report_path = self.reporter.generate_report(processing_result, self.temp_dir, config)
        
        # Verify timestamp is in filename
        filename = Path(report_path).name
        assert 'timestamped_excel_report_' in filename
        assert filename.endswith('.xlsx')
    
    def test_generate_report_invalid_path(self):
        """Test report generation with invalid output path."""
        processing_result = ProcessingResult(
            success=True,
            operation="test_operation",
            data={},
            processing_time_seconds=1.0
        )
        
        config = {'filename': 'test_report'}
        
        # Use invalid path
        invalid_path = "C:\\invalid<>path|with*invalid?chars"
        
        with pytest.raises(ValueError, match="Failed to generate Excel report"):
            self.reporter.generate_report(processing_result, invalid_path, config)
    
    def test_format_data(self):
        """Test data formatting summary."""
        data = {
            'project_status': {'project_name': 'Test Project'},
            'risks': [{'risk_id': 'R001'}, {'risk_id': 'R002'}],
            'deliverables': [{'deliverable_id': 'D001'}],
            'milestones': [{'milestone_id': 'M001'}, {'milestone_id': 'M002'}, {'milestone_id': 'M003'}],
            'stakeholders': [{'stakeholder_id': 'S001'}]
        }
        
        config = {}
        
        result = self.reporter.format_data(data, config)
        
        assert 'Project Status: Test Project' in result
        assert 'Risks: 2 items' in result
        assert 'Deliverables: 1 items' in result
        assert 'Milestones: 3 items' in result
        assert 'Stakeholders: 1 items' in result
    
    def test_format_data_empty(self):
        """Test data formatting with empty data."""
        data = {}
        config = {}
        
        result = self.reporter.format_data(data, config)
        
        assert result == "No data to format"
    
    def test_create_summary_sheet(self):
        """Test summary sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        processing_result = ProcessingResult(
            success=True,
            operation="test_operation",
            file_path=Path("test_file.xlsx"),
            data={
                'project_status': {
                    'project_name': 'Test Project',
                    'health_percentage': 85,
                    'total_risks': 5,
                    'high_priority_risks': 2
                }
            },
            processing_time_seconds=2.5,
            errors=['Test error'],
            warnings=['Test warning']
        )
        
        config = {'title': 'Test Summary Report'}
        
        self.reporter._create_summary_sheet(workbook, processing_result, config)
        
        # Verify sheet was created and is first
        assert len(workbook.sheetnames) == 2  # Default sheet + Summary
        assert workbook.sheetnames[0] == 'Summary'
        
        summary_sheet = workbook['Summary']
        assert summary_sheet['A1'].value == 'Test Summary Report'
        assert 'test_operation' in str(summary_sheet['B4'].value)
        assert 'Success' in str(summary_sheet['B5'].value)
    
    def test_create_risks_sheet(self):
        """Test risks sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        risks = [
            {
                'risk_id': 'RISK-001',
                'title': 'Budget Risk',
                'category': 'Financial',
                'priority': 'high',
                'status': 'open',
                'owner': 'PM',
                'probability': 0.7,
                'impact': 0.8,
                'risk_score': 0.56,
                'identified_date': '2024-01-15',
                'target_resolution_date': '2024-02-15',
                'mitigation_strategy': 'Monitor budget closely',
                'description': 'Risk of budget overrun'
            },
            {
                'risk_id': 'RISK-002',
                'title': 'Schedule Risk',
                'category': 'Schedule',
                'priority': 'medium',
                'status': 'mitigated',
                'owner': 'Lead Dev',
                'probability': 0.4,
                'impact': 0.6,
                'risk_score': 0.24
            }
        ]
        
        config = {}
        
        self.reporter._create_risks_sheet(workbook, risks, config)
        
        # Verify sheet was created
        assert 'Risks' in workbook.sheetnames
        
        risks_sheet = workbook['Risks']
        assert risks_sheet['A1'].value == 'Risks Analysis'
        
        # Check that data was written (headers should be in row 3)
        assert 'Risk ID' in str(risks_sheet['A3'].value)
        assert 'Title' in str(risks_sheet['B3'].value)
        
        # Check data rows
        assert 'RISK-001' in str(risks_sheet['A4'].value)
        assert 'Budget Risk' in str(risks_sheet['B4'].value)
    
    def test_create_deliverables_sheet(self):
        """Test deliverables sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        deliverables = [
            {
                'deliverable_id': 'DEL-001',
                'name': 'Requirements Document',
                'wbs_code': '1.1.1',
                'status': 'completed',
                'assigned_to': 'Analyst',
                'completion_percentage': 100.0,
                'start_date': '2024-01-01',
                'due_date': '2024-01-31',
                'estimated_effort_hours': 40.0,
                'actual_effort_hours': 45.0,
                'budget_allocated': 5000.0,
                'budget_spent': 4800.0,
                'description': 'Comprehensive requirements specification'
            }
        ]
        
        config = {}
        
        self.reporter._create_deliverables_sheet(workbook, deliverables, config)
        
        # Verify sheet was created
        assert 'Deliverables' in workbook.sheetnames
        
        deliverables_sheet = workbook['Deliverables']
        assert deliverables_sheet['A1'].value == 'Deliverables Tracking'
        
        # Check data
        assert 'DEL-001' in str(deliverables_sheet['A4'].value)
        assert 'Requirements Document' in str(deliverables_sheet['B4'].value)
    
    def test_create_milestones_sheet(self):
        """Test milestones sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        milestones = [
            {
                'milestone_id': 'MS-001',
                'name': 'Project Kickoff',
                'milestone_type': 'Start',
                'target_date': '2024-01-15',
                'actual_date': '2024-01-15',
                'status': 'completed',
                'owner': 'PM',
                'approval_required': False,
                'approver': '',
                'dependencies': ['MS-000'],
                'description': 'Official project start'
            }
        ]
        
        config = {}
        
        self.reporter._create_milestones_sheet(workbook, milestones, config)
        
        # Verify sheet was created
        assert 'Milestones' in workbook.sheetnames
        
        milestones_sheet = workbook['Milestones']
        assert milestones_sheet['A1'].value == 'Milestones Tracking'
        
        # Check data
        assert 'MS-001' in str(milestones_sheet['A4'].value)
        assert 'Project Kickoff' in str(milestones_sheet['B4'].value)
    
    def test_create_stakeholders_sheet(self):
        """Test stakeholders sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        stakeholders = [
            {
                'stakeholder_id': 'STK-001',
                'name': 'John Smith',
                'role': 'Project Sponsor',
                'organization': 'ABC Corp',
                'email': 'john.smith@example.com',
                'phone': '+1-555-0123',
                'influence': 'very_high',
                'interest': 'high',
                'engagement_priority': 'Manage Closely',
                'communication_frequency': 'Weekly',
                'preferred_communication_method': 'Email',
                'current_sentiment': 'Supportive',
                'last_contact_date': '2024-01-10',
                'next_contact_date': '2024-01-17',
                'key_concerns': ['Budget', 'Timeline'],
                'expectations': ['On-time delivery', 'Quality results']
            }
        ]
        
        config = {}
        
        self.reporter._create_stakeholders_sheet(workbook, stakeholders, config)
        
        # Verify sheet was created
        assert 'Stakeholders' in workbook.sheetnames
        
        stakeholders_sheet = workbook['Stakeholders']
        assert stakeholders_sheet['A1'].value == 'Stakeholders Analysis'
        
        # Check data
        assert 'STK-001' in str(stakeholders_sheet['A4'].value)
        assert 'John Smith' in str(stakeholders_sheet['B4'].value)
    
    def test_create_document_check_sheet(self):
        """Test document check sheet creation."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        data = {
            'found_files': [
                {
                    'filename': 'risk_register.xlsx',
                    'format': 'excel',
                    'size_bytes': 1024,
                    'is_readable': True
                },
                {
                    'filename': 'corrupted_file.xlsx',
                    'format': 'excel',
                    'size_bytes': 512,
                    'is_readable': False
                }
            ],
            'missing_files': [
                'stakeholder_register.xlsx',
                'project_charter.md'
            ]
        }
        
        config = {}
        
        self.reporter._create_document_check_sheet(workbook, data, config)
        
        # Verify sheet was created
        assert 'Document Check' in workbook.sheetnames
        
        doc_check_sheet = workbook['Document Check']
        assert doc_check_sheet['A1'].value == 'Document Check Results'
        
        # Check found files section
        assert 'Found Documents' in str(doc_check_sheet['A3'].value)
        assert 'risk_register.xlsx' in str(doc_check_sheet['A5'].value)
        
        # Check missing files section should be present
        found_missing_section = False
        for row in doc_check_sheet.iter_rows():
            for cell in row:
                if cell.value and 'Missing Documents' in str(cell.value):
                    found_missing_section = True
                    break
        assert found_missing_section
    
    def test_add_project_status_summary(self):
        """Test project status summary addition."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        status_data = {
            'project_name': 'Test Project',
            'health_percentage': 85,
            'total_risks': 10,
            'high_priority_risks': 3,
            'total_deliverables': 20,
            'completed_deliverables': 15,
            'total_milestones': 8,
            'completed_milestones': 5,
            'total_stakeholders': 12
        }
        
        self.reporter._add_project_status_summary(sheet, status_data, 5)
        
        # Check that summary was added
        assert 'Project Status Overview' in str(sheet['A5'].value)
        assert 'Test Project' in str(sheet['B6'].value)
        assert '85%' in str(sheet['B7'].value)
    
    def test_add_error_summary(self):
        """Test error summary addition."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        processing_result = ProcessingResult(
            success=False,
            operation="test_operation",
            data={},
            errors=['Error 1', 'Error 2'],
            warnings=['Warning 1'],
            processing_time_seconds=1.0
        )
        
        self.reporter._add_error_summary(sheet, processing_result, 10)
        
        # Check that error summary was added
        assert 'Processing Issues' in str(sheet['A10'].value)
        assert 'Errors:' in str(sheet['A11'].value)
        assert '• Error 1' in str(sheet['A12'].value)
        assert 'Warnings:' in str(sheet['A15'].value)
        assert '• Warning 1' in str(sheet['A16'].value)
    
    def test_write_dataframe_to_sheet(self):
        """Test DataFrame writing to sheet."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Create test DataFrame
        df = pd.DataFrame({
            'Column A': ['Value 1', 'Value 2'],
            'Column B': ['Value 3', 'Value 4'],
            'Column C': ['Value 5', 'Value 6']
        })
        
        self.reporter._write_dataframe_to_sheet(sheet, df, "Test Data")
        
        # Check title
        assert sheet['A1'].value == "Test Data"
        
        # Check headers (should be in row 3)
        assert 'Column A' in str(sheet['A3'].value)
        assert 'Column B' in str(sheet['B3'].value)
        assert 'Column C' in str(sheet['C3'].value)
        
        # Check data
        assert 'Value 1' in str(sheet['A4'].value)
        assert 'Value 3' in str(sheet['B4'].value)
    
    def test_conditional_formatting_methods(self):
        """Test conditional formatting application methods."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Set up test data for risks
        sheet['D4'] = 'Critical'
        sheet['D5'] = 'High'
        sheet['D6'] = 'Medium'
        sheet['D7'] = 'Low'
        
        # Apply formatting
        self.reporter._apply_conditional_formatting_risks(sheet, 4)
        
        # Check that cells have fill colors (we can't easily test exact colors, but we can check they're set)
        assert sheet['D4'].fill.start_color.index != '00000000'  # Not default (transparent)
        assert sheet['D5'].fill.start_color.index != '00000000'
        assert sheet['D6'].fill.start_color.index != '00000000'
        assert sheet['D7'].fill.start_color.index != '00000000'
    
    def test_auto_adjust_columns(self):
        """Test column width auto-adjustment."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Add test data with varying lengths
        sheet['A1'] = 'Short'
        sheet['A2'] = 'This is a much longer text that should affect column width'
        sheet['B1'] = 'Medium length text'
        
        self.reporter._auto_adjust_columns(sheet)
        
        # Check that column widths were adjusted
        col_a_width = sheet.column_dimensions['A'].width
        col_b_width = sheet.column_dimensions['B'].width
        
        # Column A should be wider due to the long text
        assert col_a_width > 10
        assert col_b_width > 5
        
        # Should be capped at 50
        assert col_a_width <= 50
        assert col_b_width <= 50
    
    def test_chart_creation_methods(self):
        """Test chart creation methods."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        sheet = workbook.active
        
        # Test risk priority chart
        risks = [
            {'priority': 'high'},
            {'priority': 'medium'},
            {'priority': 'high'},
            {'priority': 'low'},
            {'priority': 'critical'}
        ]
        
        self.reporter._create_risk_priority_chart(sheet, risks)
        
        # Check that data was written
        assert 'Risk Priority Distribution' in str(sheet['A1'].value)
        assert 'Low' in str(sheet['A2'].value) or 'Medium' in str(sheet['A2'].value) or 'High' in str(sheet['A2'].value)
        
        # Check that chart was added (we can't easily test chart properties, but we can check it exists)
        assert len(sheet._charts) > 0
    
    def test_get_supported_config_options(self):
        """Test getting supported configuration options."""
        options = self.reporter.get_supported_config_options()
        
        # Check base options are included
        assert 'include_timestamp' in options
        assert 'include_errors' in options
        assert 'template' in options
        
        # Check Excel-specific options
        assert 'title' in options
        assert 'filename' in options
        assert 'include_charts' in options
        assert 'apply_formatting' in options
        assert 'auto_adjust_columns' in options
        
        # Verify option structure
        assert options['title']['type'] == str
        assert options['title']['default'] == 'Project Management Analysis Report'
        assert 'description' in options['title']
        
        assert options['include_charts']['type'] == bool
        assert options['include_charts']['default'] == True
    
    def test_color_scheme(self):
        """Test color scheme definition."""
        colors = self.reporter.colors
        
        required_colors = ['header', 'success', 'warning', 'error', 'info', 'light_gray']
        for color in required_colors:
            assert color in colors
            assert colors[color].startswith('FF')  # Should be hex color with alpha
            assert len(colors[color]) == 8  # ARGB format
    
    def test_string_representations(self):
        """Test string representation methods."""
        str_repr = str(self.reporter)
        assert 'Excel Reporter (excel)' == str_repr
        
        repr_str = repr(self.reporter)
        assert 'ExcelReporter(' in repr_str
        assert "reporter_name='Excel Reporter'" in repr_str
        assert "output_format='excel'" in repr_str
        assert "file_extension='.xlsx'" in repr_str
    
    def test_empty_data_handling(self):
        """Test handling of empty data sections."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        
        # Test with empty lists
        self.reporter._create_risks_sheet(workbook, [], {})
        self.reporter._create_deliverables_sheet(workbook, [], {})
        self.reporter._create_milestones_sheet(workbook, [], {})
        self.reporter._create_stakeholders_sheet(workbook, [], {})
        
        # Should not create sheets for empty data
        sheet_names = workbook.sheetnames
        assert 'Risks' not in sheet_names
        assert 'Deliverables' not in sheet_names
        assert 'Milestones' not in sheet_names
        assert 'Stakeholders' not in sheet_names
    
    def test_add_charts_comprehensive(self):
        """Test comprehensive chart addition."""
        from openpyxl import Workbook
        
        workbook = Workbook()
        data = {
            'risks': [
                {'priority': 'high'},
                {'priority': 'medium'},
                {'priority': 'low'}
            ],
            'deliverables': [
                {'status': 'completed'},
                {'status': 'in_progress'},
                {'status': 'not_started'}
            ],
            'milestones': [
                {'status': 'completed'},
                {'status': 'upcoming'}
            ]
        }
        
        config = {}
        
        self.reporter._add_charts(workbook, data, config)
        
        # Check that charts sheet was created
        assert 'Charts' in workbook.sheetnames
        
        charts_sheet = workbook['Charts']
        
        # Check that multiple charts were added
        assert len(charts_sheet._charts) >= 2  # Should have at least risk and deliverable charts
        
        # Check that data for charts was written
        assert 'Risk Priority Distribution' in str(charts_sheet['A1'].value)
        assert 'Deliverable Status Distribution' in str(charts_sheet['A10'].value)